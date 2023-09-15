# Importe das bibliotecas 

import os
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options

#=================================================================

# Configurações do Selenium
# Caaso mude de web-drive sera necessario auterar essa região
# Recomendo utlixar o firefox por motivos de bypass de aspectos de sehurança
options = Options()
options.set_preference("general.useragent.override", "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/117.0")
driver = webdriver.Firefox(options=options)

#======================================================================

# URL do site
url = 'https://www.voudemake.com.br/'

# Abra a página da web com o Selenium
driver.get(url)

#======================================================================

# Esse trecho pode ser removido de acordo com a necessidade (ele lida com o pop-up de inicio na pagina)
driver.find_element(By.XPATH, '/html/body/data/footer/section[3]/div/div/span').click()

#======================================================================

# Abrir a planilha .xlsx
planilha = openpyxl.load_workbook('Produtos.xlsx')
planilha_ativa = planilha['Planilha2']  # Use o nome da aba 'Planilha2'

# Inicialize um contador de linha
linha_atual = 3  # Inicie em 3 para coincidir com a primeira linha de dados na planilha

#======================================================================

# Iterar sobre os valores na coluna "F" da planilha
for linha in planilha_ativa.iter_rows(min_row=3, min_col=6, max_col=6, values_only=True):
    pesquisa = linha[0]
    
    # Localize o campo de pesquisa e insira o valor
    campo_pesquisa = driver.find_element(By.XPATH, '/html/body/header/article/section/div/section/form/input')
    campo_pesquisa.clear()  # Limpar o campo de pesquisa, se houver algum texto prévio
    campo_pesquisa.send_keys(pesquisa)
    
    # Realize a pesquisa
    driver.find_element(By.XPATH, '/html/body/header/article/section/div/section/form/button').click()
    
    # Aguarde um tempo para que os resultados da pesquisa sejam exibidos
    time.sleep(10)

#======================================================================

    # Encontre todos os elementos com a classe 'department-content'
    elements = driver.find_elements(By.CLASS_NAME, 'department-content')
    
    # Verifique se a classe 'shelf__bestPrice' existe dentro dos elementos
    best_price_found = False
    for element in elements:
        best_price_elements = element.find_elements(By.CLASS_NAME, 'shelf__bestPrice')
        if best_price_elements:
            # Pegue o texto do primeiro elemento com a classe 'shelf__bestPrice'
            best_price_text = best_price_elements[0].text
            
            # Encontre a célula ao lado da célula de pesquisa atual
            # Obtendo a célula atual e avançando uma coluna à direita
            celula_pesquisa = planilha_ativa.cell(row=linha_atual, column=6)
            celula_preco = planilha_ativa.cell(row=linha_atual, column=7)
            
            # Escreva o texto do preço na célula ao lado da célula de pesquisa
            celula_preco.value = best_price_text
            
            # Encontre a imagem dentro do elemento 'department-content'
            img_element = None
            try:
                img_element = element.find_element(By.TAG_NAME, 'img')
            except:
                pass

#===========================================================================================

            if img_element:
                # Obtenha o link da imagem (atributo src)
                img_src = img_element.get_attribute('src')
                
                # Encontre a célula ao lado da célula de pesquisa atual
                # Obtendo a célula atual e avançando uma coluna à direita
                celula_imagem = planilha_ativa.cell(row=linha_atual, column=4)
                
                # Escreva o link da imagem na célula à direita do preço
                celula_imagem.value = img_src
            
            best_price_found = True
            break  # Pare a busca após encontrar a primeira ocorrência

#========================================================================================
    
    if not best_price_found:
        # Se 'shelf__bestPrice' não foi encontrado, procure 'shelf__out-of-stock'
        for element in elements:
            out_of_stock_elements = element.find_elements(By.CLASS_NAME, 'shelf__out-of-stock')
            if out_of_stock_elements:
                # Pegue o texto do primeiro elemento com a classe 'shelf__out-of-stock'
                out_of_stock_text = out_of_stock_elements[0].text
                
                # Encontre a célula ao lado da célula de pesquisa atual
                # Obtendo a célula atual e avançando uma coluna à direita
                celula_pesquisa = planilha_ativa.cell(row=linha_atual, column=6)
                celula_preco = planilha_ativa.cell(row=linha_atual, column=7)
                
                # Escreva o texto 'Fora de Estoque' na célula ao lado da célula de pesquisa
                celula_preco.value = 'Fora de Estoque'
                
                break  # Pare a busca após encontrar a primeira ocorrência

#======================================================================
    
    # Salve a planilha após cada iteração
    planilha.save('Produtos.xlsx')
    
    # Incremente o contador de linha
    linha_atual += 1

#======================================================================

# Salve a planilha final após concluir todas as pesquisas
planilha.save('Produtos.xlsx')

# Feche o navegador quando terminar
driver.quit()
